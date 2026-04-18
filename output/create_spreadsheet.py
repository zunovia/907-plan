import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# --- Style definitions ---
title_font = Font(name="Meiryo", size=14, bold=True)
header_font = Font(name="Meiryo", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
sub_header_font = Font(name="Meiryo", size=11, bold=True)
sub_header_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
body_font = Font(name="Meiryo", size=10)
wrap_align = Alignment(wrap_text=True, vertical="top")
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
accent_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
red_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")


def apply_header(ws, row, cols, widths=None):
    for i, col in enumerate(cols, 1):
        cell = ws.cell(row=row, column=i, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w


def add_row(ws, row, values, font=body_font, fill=None):
    for i, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=i, value=v)
        cell.font = font
        cell.alignment = wrap_align
        cell.border = thin_border
        if fill:
            cell.fill = fill


# ============================================================
# Sheet 1: エグゼクティブサマリー
# ============================================================
ws1 = wb.active
ws1.title = "エグゼクティブサマリー"
ws1.cell(row=1, column=1, value="中部リサイクル運動市民の会 — 新規事業・収益拡大戦略 最終報告書").font = title_font
ws1.merge_cells("A1:D1")

r = 3
info = [
    ("議題", "中部リサイクル運動市民の会の新規事業・収益拡大戦略"),
    ("実施日", "2026-04-18"),
    ("参加者", "CEO, CTO, CFO, CMO, 新規事業企画, 品質保証・リスク管理"),
    ("ラウンド数", "2ラウンド（Round 2で合意形成）"),
    ("対象組織", "NPO法人 中部リサイクル運動市民の会 + 一般社団法人 サーコミュニケーション"),
]
for label, val in info:
    ws1.cell(row=r, column=1, value=label).font = Font(name="Meiryo", size=10, bold=True)
    ws1.cell(row=r, column=1).fill = sub_header_fill
    ws1.cell(row=r, column=1).border = thin_border
    ws1.cell(row=r, column=2, value=val).font = body_font
    ws1.cell(row=r, column=2).border = thin_border
    ws1.cell(row=r, column=2).alignment = wrap_align
    r += 1

r += 1
ws1.cell(row=r, column=1, value="エグゼクティブサマリー").font = Font(name="Meiryo", size=12, bold=True)
r += 1
summary = (
    "NPO法人「中部リサイクル運動市民の会」の年間売上1億円達成は、NPO単体ではなく関連法人「サーコミュニケーション」との"
    "グループ戦略として3年計画で実現可能と判断。最優先施策はEC展開による既存リサイクル事業のオンライン拡張であり、"
    "初期投資10万円以下でMVP検証を即時開始できる。NPO法の収益事業制約への対応として、収益性の高い事業は"
    "サーコミュニケーションを事業主体とする「二法人戦略」を全員一致で採択。"
)
ws1.cell(row=r, column=1, value=summary).font = body_font
ws1.cell(row=r, column=1).alignment = wrap_align
ws1.merge_cells(f"A{r}:D{r}")
ws1.row_dimensions[r].height = 80

ws1.column_dimensions["A"].width = 18
ws1.column_dimensions["B"].width = 60
ws1.column_dimensions["C"].width = 20
ws1.column_dimensions["D"].width = 20

# ============================================================
# Sheet 2: 総合評価
# ============================================================
ws2 = wb.create_sheet("総合評価")
ws2.cell(row=1, column=1, value="総合評価スコア").font = title_font
apply_header(ws2, 3, ["参加者", "立場", "総合評価"], [20, 35, 15])
evals = [
    ("biz-ceo (CEO)", "条件付き賛成", "★★★★☆"),
    ("biz-cto (CTO)", "条件付き賛成", "★★★★☆"),
    ("biz-cfo (CFO)", "要検討→条件付き賛成（Round 2で修正）", "★★★☆☆"),
    ("biz-marketing (CMO)", "条件付き賛成", "★★★★☆"),
    ("biz-newbiz (新規事業企画)", "賛成", "★★★★☆"),
    ("biz-qa (品質保証・リスク管理)", "条件付き賛成", "★★★☆☆"),
]
for i, row_data in enumerate(evals, 4):
    add_row(ws2, i, row_data)

# ============================================================
# Sheet 3: 合意事項
# ============================================================
ws3 = wb.create_sheet("合意事項")
ws3.cell(row=1, column=1, value="合意事項（6項目）").font = title_font
apply_header(ws3, 3, ["#", "項目", "詳細"], [5, 25, 80])
agreements = [
    (1, "二法人戦略の採択", "NPO本体はリサイクルショップ・環境教育・地域活動に集中。収益性の高い新規事業（法人コンサル・DX外販）はサーコミュニケーションを事業主体とする。"),
    (2, "EC展開を最優先施策", "BASEで高単価商品（家電・ブランド品）30品からスタート。初期投資10万円以下。月商30万円を3ヶ月で達成できれば本格展開。"),
    (3, "段階的投資アプローチ（3 Tier制）", "Tier1: EC展開+既存店舗最大化（即時）、Tier2: 出張買取+法人コンサル（6ヶ月以内）、Tier3: DX外販+研修拡大（1年後）"),
    (4, "法人コンサルはセミナー起点", "名古屋市の環境経営セミナー登壇→関心企業の問い合わせ→提案。サーコミュニケーションが事業主体。"),
    (5, "助成金の積極活用", "環境省「循環経済実現促進事業」、経産省「IT導入補助金」等で年間200-500万円の獲得を目指す。"),
    (6, "品質基準・返品ポリシーの策定", "リサイクル品のグレーディングシステム（A/B/Cランク）導入。返品率5%以下をKPIとする。"),
]
for i, row_data in enumerate(agreements, 4):
    add_row(ws3, i, row_data)
    ws3.row_dimensions[i].height = 40

# ============================================================
# Sheet 4: 売上ロードマップ
# ============================================================
ws4 = wb.create_sheet("売上ロードマップ")
ws4.cell(row=1, column=1, value="グループ売上構成（3年計画）").font = title_font
apply_header(ws4, 3, ["事業", "主体", "Year 1", "Year 2", "Year 3"], [30, 15, 15, 15, 15])
sales = [
    ("リサイクルショップ（3店舗）", "NPO", 70000000, 72000000, 75000000),
    ("EC売上", "NPO", 2000000, 6000000, 10000000),
    ("JICA研修・名古屋市受託", "NPO", 4000000, 3500000, 3000000),
    ("出張買取", "NPO", 0, 1500000, 3000000),
    ("会費・カンパ・助成金", "NPO", 2000000, 3000000, 4000000),
    ("法人コンサル", "サーコミュ", 1000000, 3000000, 6000000),
    ("DX外販", "サーコミュ", 0, 1000000, 4000000),
    ("環境教育・研修（国内）", "NPO/サーコミュ", 1000000, 2000000, 3000000),
]
yen_fmt = '#,##0"円"'
for i, (name, entity, y1, y2, y3) in enumerate(sales, 4):
    add_row(ws4, i, [name, entity, y1, y2, y3])
    for col in [3, 4, 5]:
        ws4.cell(row=i, column=col).number_format = yen_fmt

total_row = 4 + len(sales)
add_row(ws4, total_row, ["グループ合計", "", 80000000, 92000000, 108000000],
        font=Font(name="Meiryo", size=10, bold=True), fill=accent_fill)
for col in [3, 4, 5]:
    ws4.cell(row=total_row, column=col).number_format = yen_fmt

# Milestones
r = total_row + 3
ws4.cell(row=r, column=1, value="マイルストーン").font = Font(name="Meiryo", size=12, bold=True)
r += 1
apply_header(ws4, r, ["マイルストーン", "時期", "売上目標", "主要アクション", ""], [30, 18, 18, 50, 5])
milestones = [
    ("EC MVP開始", "2026年5月", "-", "BASE開設、高単価商品30品出品"),
    ("EC本格展開判断", "2026年8月", "EC月商30万", "月商30万達成でShopify移行検討"),
    ("SNS・PR基盤構築", "2026年7月", "-", "Instagram/LINE開設、地元メディアPR"),
    ("品質基準策定", "2026年6月", "-", "グレーディングシステム・返品ポリシー"),
    ("法人コンサルMVP", "2026年10月", "-", "環境セミナー登壇、3社ヒアリング"),
    ("出張買取開始", "2027年1月", "月15万", "名古屋市内限定でテスト"),
    ("Year 1着地", "2027年3月", "8,000万", "既存事業最適化+EC立ち上げ完了"),
    ("DX外販MVP", "2027年10月", "-", "3NPOに経理自動化パッケージ無償提供"),
    ("EC月商100万達成", "2027年12月", "EC月商100万", "商品数500品、AI画像処理導入"),
    ("Year 2着地", "2028年3月", "9,200万", "新規事業が収益貢献開始"),
    ("法人コンサル10社", "2028年12月", "月50万", "サーコミュの主力事業化"),
    ("DX外販本格化", "2029年1月", "-", "有償化、年間5件受注目標"),
    ("Year 3着地", "2029年3月", "10,800万", "グループ売上1億円達成"),
]
for i, row_data in enumerate(milestones, r + 1):
    fill = accent_fill if "Year" in row_data[0] else None
    add_row(ws4, i, row_data, fill=fill)

# ============================================================
# Sheet 5: リスク分析
# ============================================================
ws5 = wb.create_sheet("リスク分析")
ws5.cell(row=1, column=1, value="リスクと軽減策").font = title_font
apply_header(ws5, 3, ["#", "リスク", "影響度", "発生確率", "対応策"], [5, 30, 10, 10, 55])
risks = [
    (1, "NPO法の収益事業比率超過", "高", "中", "収益事業はサーコミュに移管。顧問弁護士・所轄庁に事前確認"),
    (2, "リサイクル品のEC品質クレーム", "中", "高", "グレーディングシステム（A/B/Cランク）導入。返品率5%以下KPI"),
    (3, "リサイクラー人材不足・高齢化", "高", "中", "若年層リクルーティング、大学・専門学校との連携、デジタルツール導入"),
    (4, "競合（セカンドストリート等）の攻勢", "中", "中", "NPOブランド・環境貢献の差別化。地域密着・コミュニティ価値の訴求"),
    (5, "サーコミュニケーションへの技術依存", "中", "中", "NPO側にIT担当1名を育成。日常運用の内製化"),
    (6, "NPO-サーコミュ間の利益相反", "中", "低", "理事会での取引承認制度、外部監査の活用"),
    (7, "個人情報漏洩（CRM・EC）", "高", "低", "セキュリティポリシー策定、SSL/TLS、外部決済サービス利用"),
    (8, "名古屋市との関係変化", "高", "低", "定期的な情報共有・連携強化、受託事業の品質維持"),
    (9, "助成金の不採択", "中", "中", "複数の助成金に並行申請。助成金なしでも事業継続可能な計画設計"),
    (10, "移転価格税務リスク", "中", "低", "NPO-サーコミュ間取引を市場価格ベースに設定。税理士との事前相談"),
]
for i, row_data in enumerate(risks, 4):
    impact = row_data[2]
    fill = red_fill if impact == "高" else None
    add_row(ws5, i, row_data, fill=fill)
    ws5.row_dimensions[i].height = 30

# 撤退基準
r = 4 + len(risks) + 2
ws5.cell(row=r, column=1, value="撤退基準").font = Font(name="Meiryo", size=12, bold=True)
r += 1
apply_header(ws5, r, ["事業", "撤退基準", "判断時期", "", ""], [20, 40, 18, 10, 10])
exits = [
    ("EC事業", "6ヶ月連続で月商30万円未達", "2026年11月"),
    ("法人コンサル", "12ヶ月で契約3社未満", "2027年10月"),
    ("DX外販", "MVP提供後12ヶ月で有償契約ゼロ", "2028年10月"),
    ("出張買取", "6ヶ月連続で月売上10万円未満", "2027年7月"),
]
for i, row_data in enumerate(exits, r + 1):
    add_row(ws5, i, row_data)

# ============================================================
# Sheet 6: ネクストステップ
# ============================================================
ws6 = wb.create_sheet("ネクストステップ")
ws6.cell(row=1, column=1, value="ネクストステップ").font = title_font

r = 3
ws6.cell(row=r, column=1, value="即時実行（2026年4月-5月）").font = Font(name="Meiryo", size=12, bold=True)
r += 1
apply_header(ws6, r, ["#", "アクション", "担当", "期限"], [5, 50, 20, 18])
immediate = [
    (1, "NPO法の収益事業制約について顧問弁護士に相談", "CEO/総務", "2026年4月末"),
    (2, "古物商許可のEC対応（URL届出）確認", "総務", "2026年4月末"),
    (3, "BASEアカウント開設、高単価商品30品の選定・撮影・出品", "店舗責任者+CTO", "2026年5月中旬"),
    (4, "リサイクル品グレーディング基準の策定", "QA/店舗責任者", "2026年5月末"),
    (5, "Instagram/LINE公式アカウント開設", "CMO/総務", "2026年5月中旬"),
    (6, "サーコミュニケーションとの役割分担・取引条件の協議", "CEO", "2026年5月末"),
]
for i, row_data in enumerate(immediate, r + 1):
    add_row(ws6, i, row_data)

r = r + 1 + len(immediate) + 1
ws6.cell(row=r, column=1, value="短期（2026年6月-12月）").font = Font(name="Meiryo", size=12, bold=True)
r += 1
apply_header(ws6, r, ["#", "アクション", "担当", "期限"], [5, 50, 20, 18])
short_term = [
    (7, "Googleマイビジネス最適化（3店舗全て）", "CMO", "2026年6月末"),
    (8, "地元メディア（中日新聞・CBC等）へのPR素材提供", "CMO", "2026年7月"),
    (9, "環境省・経産省の助成金申請", "CFO/総務", "公募に合わせて"),
    (10, "名古屋市環境経営セミナーでの登壇調整", "CEO/newbiz", "2026年9月"),
    (11, "法人コンサルの営業資料・サービス設計", "サーコミュ", "2026年10月"),
    (12, "AI画像処理ステーションの開発", "CTO/サーコミュ", "2026年12月"),
    (13, "個人情報保護ポリシー・返品ポリシーの策定・公開", "QA/総務", "2026年6月末"),
]
for i, row_data in enumerate(short_term, r + 1):
    add_row(ws6, i, row_data)

r = r + 1 + len(short_term) + 1
ws6.cell(row=r, column=1, value="中期（2027年1月-2028年3月）").font = Font(name="Meiryo", size=12, bold=True)
r += 1
apply_header(ws6, r, ["#", "アクション", "担当", "期限"], [5, 50, 20, 18])
mid_term = [
    (14, "出張買取サービスの名古屋市内テスト開始", "店舗/物流", "2027年1月"),
    (15, "EC本格展開（Shopify移行検討、商品数拡大）", "CTO", "2027年6月"),
    (16, "CRM導入（顧客管理・リピート促進）", "CTO/CMO", "2027年6月"),
    (17, "法人コンサル3社契約達成", "サーコミュ", "2027年10月"),
    (18, "DX外販MVPの無償提供開始（3NPO）", "サーコミュ/CTO", "2027年10月"),
    (19, "NPO側IT担当者の育成", "CTO", "2028年3月"),
]
for i, row_data in enumerate(mid_term, r + 1):
    add_row(ws6, i, row_data)

# ============================================================
# Sheet 7: 投資計画・役割分担
# ============================================================
ws7 = wb.create_sheet("投資計画・役割分担")
ws7.cell(row=1, column=1, value="投資計画サマリー").font = title_font
apply_header(ws7, 3, ["投資項目", "金額", "時期", "回収見込み"], [30, 18, 15, 40])
investments = [
    ("EC構築（BASE→Shopify）", "10万→50万", "Year 1", "EC月商30万で6ヶ月回収"),
    ("AI画像処理ステーション", "30万", "Year 1", "出品効率化で間接回収"),
    ("CRM導入", "50万", "Year 1-2", "リピート率向上で12ヶ月回収"),
    ("SNS/PR/マーケティング", "100万（累計）", "Year 1-3", "来店増・EC売上で回収"),
    ("法人コンサル営業資料", "50万", "Year 1", "3社契約で6ヶ月回収"),
    ("セキュリティ・コンプライアンス対応", "20万", "Year 1", "リスク回避（間接効果）"),
    ("DX外販パッケージ開発", "100万", "Year 2", "5件受注で12ヶ月回収"),
]
for i, row_data in enumerate(investments, 4):
    add_row(ws7, i, row_data)
total_r = 4 + len(investments)
add_row(ws7, total_r, ["合計", "約400万", "3年間", ""], font=Font(name="Meiryo", size=10, bold=True), fill=accent_fill)

# 役割分担
r = total_r + 3
ws7.cell(row=r, column=1, value="NPO法人とサーコミュニケーションの役割分担").font = Font(name="Meiryo", size=12, bold=True)
r += 1
apply_header(ws7, r, ["領域", "NPO本体", "サーコミュニケーション", ""], [25, 30, 30, 5])
roles = [
    ("リサイクルショップ運営", "主体", "技術支援（POS・在庫管理）"),
    ("EC事業", "主体（販売）", "技術支援（EC基盤・AI画像処理）"),
    ("出張買取", "主体", "-"),
    ("JICA研修・名古屋市受託", "主体", "-"),
    ("環境教育・研修", "共同", "共同（オンライン基盤）"),
    ("法人コンサル", "-", "主体"),
    ("DX外販", "ノウハウ提供", "主体（開発・販売）"),
    ("経理・業務自動化", "利用者", "開発・運用"),
    ("ブランド", "環境貢献・市民活動", "ビジネス・技術力"),
]
for i, row_data in enumerate(roles, r + 1):
    add_row(ws7, i, row_data)

# ============================================================
# Sheet 8: 専門家見解
# ============================================================
ws8 = wb.create_sheet("専門家見解")
ws8.cell(row=1, column=1, value="各専門家の見解サマリー").font = title_font
apply_header(ws8, 3, ["専門家", "見解サマリー"], [22, 90])
experts = [
    ("biz-ceo (CEO)", "既存インフラ活用と二法人戦略を軸に段階的事業拡大を推奨。NPO+サーコミュのグループ経営視点で意思決定すべき。地元メディアPRによる低コスト認知拡大と名古屋商工会議所等との連携でリード獲得を提案。"),
    ("biz-cto (CTO)", "既存技術基盤（Python自動化・OCR・AI）がNPO業界では突出。EC展開時のボトルネックはAI画像処理ステーションで解決可能（開発コスト30万円）。長期的にはDXノウハウの外販がスケーラブルな収益源。サーコミュの技術者1名追加が前提条件。"),
    ("biz-cfo (CFO)", "3年間のP&L試算を提示。Year1:8,000万、Year2:9,000万、Year3:1億円。助成金活用（年間200-500万円）の重要性を強調。NPO→サーコミュ間の移転価格リスクへの注意を喚起。"),
    ("biz-marketing (CMO)", "「環境貢献×お得な買い物」のバリュープロポジションを軸に、Instagram/LINE/Googleマイビジネス活用のデジタルマーケティング戦略を提案。法人向けは「サーコミュ」ブランド、個人向けは「NPO」ブランドの使い分けを提案。"),
    ("biz-newbiz (新規事業企画)", "5つの新規事業候補を優先順位付きで提案。(1)法人コンサル(A) (2)ECショップ(A) (3)DX外販(B+) (4)環境教育(B) (5)出張買取(B)。EC展開のMVPを「BASE上で高単価商品30品、月商30万円を3ヶ月」と定義。"),
    ("biz-qa (品質保証・リスク管理)", "NPO法の収益事業制約、古物営業法のEC対応等のコンプライアンスチェックリストを提示。二法人戦略の利益相反リスクを指摘。各事業の撤退基準を明確化（EC:6ヶ月で月商30万未達→縮小等）。"),
]
for i, row_data in enumerate(experts, 4):
    add_row(ws8, i, row_data)
    ws8.row_dimensions[i].height = 65

# Save
output_path = r"C:\02.Claude_Agent\907-plan\output\roundtable-report.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
